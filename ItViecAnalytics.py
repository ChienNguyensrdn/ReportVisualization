class JobDetail:
    def __init__(self,company,location,linkDetail,jobTitle,jobTag,time):
        self._company=company
        self._location=location
        self._linkDetail=linkDetail
        self._jobTitle=jobTitle
        self._jobTag=jobTag
        self._time=time
from matplotlib import colors
import requests
from bs4 import BeautifulSoup
import matplotlib.pyplot as plt
import seaborn as sns
from collections import Counter
import numpy as np
from datetime import date,timedelta
import re
import openpyxl
import csv

listJob = []
locations=['ho-chi-minh-hcm','ha-noi','da-nang']
header = ['company', 'location', 'linkDetail', 'jobTitle','jobTag','time']
with open('Data/DataItviec.csv', 'w') as f:
    writer = csv.writer(f)
    # write the header
    writer.writerow(header)
    for loc in locations:
        rangeTo =0
        if loc=='ho-chi-minh-hcm':
            rangeTo=53 #53
        elif loc=='ha-noi':
            rangeTo=26 #26
        else:
            rangeTo=5
        for i in range(1,rangeTo):
            url = requests.get("https://itviec.com/viec-lam-it/"+loc+"?locale=vi&page="+str(i)+"&source=search_job")
            html = url.text
            # print(htmltext)
            soup = BeautifulSoup(html, features="html.parser")
            job_content =soup.find_all('div',class_='job_content')
            for item in job_content:
                linkDetail =item.find_all('a')[0]['href']
                print('link detail: '+linkDetail)
                print('Logo company: '+ item.find_all('img', alt=True)[0]['alt'])
                jobTitle=item.find_all('h2',class_='title')[0]
                jobTag =item.find_all('div',class_='tag-list')[0]
                print('jobTitle: '+ jobTitle.find_all('a')[0].text)
                print('Job tag: '+ jobTag.find_all('a')[0].text )
                salaryRank=item.find_all('div', class_ = 'svg-icon__text')[0]
                print('Salary Rank '+salaryRank.find_all('a')[0].text)
                jobTime =item.find_all('div', class_ = 'distance-time-job-posted')[0]
                print('Job Time: '+jobTime.find_all('span')[0].text)
                jobTimeDate=date.today()
                if jobTime.find_all('span')[0].text.upper().find('GIỜ') >0 :
                    jobTimeDate=date.today()
                else:
                    number= re.findall(r'\d', jobTime.find_all('span')[0].text)  
                    jobTimeDate = date.today()+timedelta(days=int(number[0])*-1) if number[0].isnumeric()  else date.today()
                print(jobTimeDate)
                listJob.append(
                    JobDetail(
                        item.find_all('img', alt=True)[0]['alt'], 
                        loc,
                        linkDetail,
                        jobTitle.find_all('a')[0].text.strip(),
                        jobTag.find_all('a')[0].text.strip(),
                        jobTimeDate
                ))
                data=[ 
                    item.find_all('img', alt=True)[0]['alt'],
                    loc,
                    linkDetail,
                    jobTitle.find_all('a')[0].text.strip(),
                    jobTag.find_all('a')[0].text.strip(),
                    jobTimeDate
                ]
                writer.writerow(data)


# wb = openpyxl.load_workbook('Data/Book1.xlsx')
# sheet = wb['Sheet1']
# idxRow=2
# for item in listJob:
#     try:
#         sheet.cell(row=idxRow, column=1, value=item._company)
#         sheet.cell(row=idxRow, column=2, value=item._linkDetail)
#         sheet.cell(row=idxRow, column=3, value=item._jobTitle)
#         sheet.cell(row=idxRow, column=4, value=item._jobTag)
#         sheet.cell(row=idxRow, column=5, value=item._time)
#         sheet.cell(row=idxRow, column=6, value=item._location)
#         idxRow+=1
#     except:
#         continue
    
# wb.save('Data/Book1.xlsx')

# jobTags =[rec._jobTag for rec in listJob]
# counts =Counter(jobTags)
# t=counts.most_common(20)
# keys=[]
# values=[]
# for item in t:
#     keys.append(item[0])
#     values.append(item[1])
# sns.barplot(values , keys)
# plt.ylabel('Job tag')
# plt.xlabel('target')
# jobLocation=[rec._location for rec in listJob]
# counts =Counter(jobLocation)
# t=counts.most_common(3)
# keys=[]
# values=[]
# for item in t:
#     keys.append(item[0])
#     values.append(item[1])
# sns.barplot(keys,values)
# plt.ylabel('Job tag')
# plt.xlabel('target')

jobDay = [rec._time for rec in listJob] 
counts =Counter(jobDay)
t=counts.most_common(30)
keys=[]
values=[]
for item in t:
    keys.append(item[0])#.strftime('%b')+'-'+item[0].strftime('%d'))
    values.append(item[1])
sns.lineplot(keys,values)
plt.ylabel('Job count')
plt.xlabel('Time line')
plt.show()